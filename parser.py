from openpyxl import load_workbook
import imaplib
import email
from email.header import decode_header
import base64
from bs4 import BeautifulSoup
import re
from io import BytesIO
from work_pochta import get_attachments

# библиотеки для парсинга, для подключения к почте

#book = load_workbook(filename=BytesIO(get_attachments()))

book = load_workbook(filename="C:/Users/Deserag/Downloads/Telegram Desktop/10%2013.03-18.03.xlsx")
#выбор нужного файла из почты

sheet = book['Колледж ВятГУ']
pattern = r"Группа .{2,3}к"
for i in range(65, 91):
    val = sheet[chr(i) + '24'].value
    if val and val[0] == " ":
        val = val[1:len(val)]
    if val and re.match(pattern, val):
        begin_value = i
        print(sheet[chr(i) + '24'].value)
        break
simvol = chr(begin_value)
kolich = 0

def change_sim(letter):
    if len(letter) < 2:
        if ord(letter) < 90:
            letter = chr(ord(letter) + 1)
        else:
            letter = "AA"
    else:
        if ord(letter[-1]) < 90:
            letter = letter[0:len(letter) - 1] + chr(ord(letter[-1]) + 1)
        else:
            letter = chr(ord(letter[-2]) + 1) + "A"
    return letter

pars = ['1 пара', '2 пара', '3 пара', '4 пара', '5 пара', '6 пара', '7 пара'] #номера пар
days = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота', 'воскресенье'] #номер дня недели
para = ['8:20-9:50','10:00-11:30','11:45-13:15','14:00-15:30','15:45-17:15','17:20-18:50','18:55-20:25'] #время звонков по будням
para_subbot = ['8:20-9:50','10:00-11:30','11:45-13:15','13:20-14:50','14:55-16:25','16:30-18:00'] #время звонков по субботам
done_group = 0
groups_with_data = []
while True:
    if kolich >= 10:
        break
    val = sheet[simvol + '24'].value
    if val:
        kolich = 0
        x = 26

        for i in range(41):
            if days[i // 7] == 'суббота':
                groups_with_data.append([val, pars[i % 7], para_subbot[i % 7], days[i // 7], sheet[simvol + str(x)].value])

            else:
                groups_with_data.append([val, pars[i % 7],para[i % 7], days[i // 7], sheet[simvol + str(x)].value])

            x += 1
        x = 26
        simvol = change_sim(simvol)
        for i in range(41):
            groups_with_data[done_group].append(sheet[simvol + str(x)].value)
            done_group += 1
            x += 1
        done_group -= 41
        x = 26
        simvol = change_sim(simvol)

        for i in range(41):
            groups_with_data[done_group].append(sheet[simvol + str(x)].value)
            done_group += 1
            x += 1
        done_group -= 41
        x = 26
        simvol = change_sim(simvol)

        for i in range(41):
            groups_with_data[done_group].append(sheet[simvol + str(x)].value)
            done_group += 1
            x += 1
        simvol = change_sim(simvol)

    else:
        simvol = change_sim(simvol)

        kolich += 1
for i in groups_with_data:
    print(i)
