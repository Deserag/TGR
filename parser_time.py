from openpyxl import load_workbook
import imaplib
import email
from email.header import decode_header
import base64
from bs4 import BeautifulSoup
import re
from io import BytesIO
from work_pochta import get_attachments

# библиотеки для парсинга, для подключения к почте


import psycopg2

#библиотека для работы с postgresql


# установление соединения с базой данных
conn = psycopg2.connect(dbname='BotDBtest', user='postgres', password='danilworld1', host='127.0.0.1', port='5432')

# создание объекта курсора
cur = conn.cursor()






#book = load_workbook(filename=BytesIO(get_attachments()))

book = load_workbook(filename="13_03_04-08_04.xlsx")
#выбор нужного файла из почты

sheet = book['Колледж ВятГУ']
pattern = r"Группа .{2,3}к"
for i in range(65, 91):
    val = sheet[chr(i) + '24'].value
    if val and val[0] == " ":
        val = val[1:len(val)]
    if val and re.match(pattern, val):
        begin_value = i
        print(sheet[chr(i) + '24'].value)
        break
simvol = chr(begin_value)
kolich = 0

def change_sim(letter):
    if len(letter) < 2:
        if ord(letter) < 90:
            letter = chr(ord(letter) + 1)
        else:
            letter = "AA"
    else:
        if ord(letter[-1]) < 90:
            letter = letter[0:len(letter) - 1] + chr(ord(letter[-1]) + 1)
        else:
            letter = chr(ord(letter[-2]) + 1) + "A"
    return letter

pars = ['1 пара', '2 пара', '3 пара', '4 пара', '5 пара', '6 пара', '7 пара']
days = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота', '7 пара']
para = ['8:20-9:50','10:00-11:30','11:45-13:15','14:00-15:30','15:45-17:15','17:20-18:50','18:55-20:25']
para_subbot = ['8:20-9:50','10:00-11:30','11:45-13:15','13:20-14:50','14:55-16:25','16:30-18:00']
done_group = 0
groups_with_data = []
while True:
    if kolich >= 10:
        break
    val = sheet[simvol + '24'].value
    if val:
        kolich = 0
        x = 26

        for i in range(41):
            if days[i // 7] == 'суббота':
                groups_with_data.append([val, pars[i % 7], para_subbot[i % 7], days[i // 7], sheet[simvol + str(x)].value])
            else:
                groups_with_data.append([val, pars[i % 7],para[i % 7], days[i // 7], sheet[simvol + str(x)].value])
            x += 1
        x = 26
        simvol = change_sim(simvol)
        for i in range(41):
            groups_with_data[done_group].append(sheet[simvol + str(x)].value)
            done_group += 1
            x += 1
        done_group -= 41
        x = 26
        simvol = change_sim(simvol)

        for i in range(41):
            groups_with_data[done_group].append(sheet[simvol + str(x)].value)
            done_group += 1
            x += 1
        done_group -= 41
        x = 26
        simvol = change_sim(simvol)

        for i in range(41):
            groups_with_data[done_group].append(sheet[simvol + str(x)].value)
            done_group += 1
            x += 1
        simvol = change_sim(simvol)

    else:
        simvol = change_sim(simvol)

        kolich += 1



#чистим пробелы


groups_with_data = [data.strip() if isinstance(data, str) else data for data in groups_with_data]

#чистим от группы



for i in range(len(groups_with_data)):
    groups_with_data[i][0] = groups_with_data[i][0].replace('Группа ', '').strip()



#делаем копии строк для групп


groups_with_data_new = []

for lesson in groups_with_data:
    if '\n' in lesson[0]:
        groups = lesson[0].split('\n')
        for group in groups:
            new_lesson = [group] + lesson[1:]
            groups_with_data_new.append(new_lesson)
    else:
        groups_with_data_new.append(lesson)



groups_with_data = groups_with_data_new




#вывод для просмотра

for i in groups_with_data:
    print(i)




#print(groups_with_data)


#запрос на удаление всех записей из таблицы lecture

#cur.execute("DELETE FROM lecture")


cur.execute("TRUNCATE TABLE lecture RESTART IDENTITY;")     #для того чтобы сбрасывался первичный ключ





#заполнения таблицы lecture новыми данными

conn.commit()


# выполнение цикла для добавления данных в таблицу
for info in groups_with_data:
        cur.execute("INSERT INTO lecture (day, subject, teacher, classroom, group_id, lecture_type, lecture_time) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (info[3], info[4], info[6], info[7], info[0], info[5], info[2]))

# сохранение изменений и закрытие соединения с базой данных
conn.commit()
conn.close()


#было использовано для тестов



# запрос на выборку данных из таблицы

#group_name = input("Введите название группы: ")
#day_of_week = input("Введите день недели: ")
#cur.execute("SELECT day, subject, teacher, classroom, lecture_type, lecture_time FROM lecture WHERE group_id ILIKE %s AND day = %s", (f"{group_name}%", day_of_week))

# получение результата и сохранение его в список
#result_list = cur.fetchall()

# вывод списка
#for row in result_list:
    #print(row)


#result_list2 = []
#for row in result_list:
    #if all(row[1:]):  # проверка, что все поля кроме поля дня недели не пустые
        #result_list2.append(row)

#for row in result_list2:
    #print(row)


#text = ""
#for row in result_list2:
    #text += f"{row[5]} / {row[1]} /{row[4]} / {row[2]} / {row[3]}\n"
#print(text)



#day_text = text






